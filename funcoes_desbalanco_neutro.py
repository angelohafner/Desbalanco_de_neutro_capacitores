
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule


# %%
def destaca_maiores_que_nominal(planilha='Ia1', aquivo='correntes.xlsx', valor_nominal=10.1):
    wb = load_workbook(aquivo)
    ws = wb[planilha]
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    dxf = DifferentialStyle(fill=red_fill)
    rule = ColorScaleRule(start_type="min", start_color="FFFF99",
                          end_type="max", end_color="FF0000")
    ws.conditional_formatting.add("A1:Z26", rule)
    wb.save(aquivo)


# %%
def matriz_complexa_para_polar(matriz_complexa):
    magnitude = np.abs(matriz_complexa)
    angulo = np.angle(matriz_complexa)

    # Gera a matriz de strings em formato polar
    matriz_polar = np.empty(matriz_complexa.shape, dtype=object)

    for i in range(matriz_complexa.shape[0]):
        for j in range(matriz_complexa.shape[1]):
            matriz_polar[i, j] = f'{magnitude[i, j]:.2f}∠{np.degrees(angulo[i, j]):.2f}°'

    return matriz_polar

# %%
def load_excel_to_numpy(filename):
    wb = load_workbook(filename=filename, read_only=True)
    sheets = ['A', 'B', 'C']
    data = {}

    for sheet in sheets:
        ws = wb[sheet]
        data[sheet] = pd.DataFrame(ws.values)

    return {k: v.to_numpy() for k, v in data.items()}

# %%
def matriz_fases_ramos(nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int, cap_interna, des_int):
    matriz = np.ones((3, nr_lin_ext * nr_lin_int, 2 * nr_col_ext * nr_col_int))
    cores = ['FFFFCC', 'CCFFCC', 'CCCCFF', 'FFCCFF']

    wb = Workbook()
    ws = wb.active
    ws.title = "A"

    for fase in ['A', 'B', 'C']:
        if fase == 'A':
            ws = wb.active
            fase_cont = 0
        else:
            ws = wb.create_sheet(title=fase)
            fase_cont = fase_cont + 1

        for i_ext in range(nr_lin_ext):
            for j_ext in range(2 * nr_col_ext):
                cor = cores[(i_ext + j_ext) % len(cores)]
                fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
                for i_int in range(nr_lin_int):
                    for j_int in range(nr_col_int):
                        ii = i_ext * nr_lin_int + i_int
                        jj = j_ext * nr_col_int + j_int
                        valor = cap_interna * np.random.uniform(1 - des_int, 1 + des_int)
                        matriz[fase_cont, ii, jj] = valor
                        ws.cell(row=ii + 1, column=jj + 1, value=valor)
                        ws.cell(row=ii + 1, column=jj + 1).fill = fill
                        ws.cell(row=ii + 1, column=jj + 1).number_format = '##0.0'
                        ws.column_dimensions[get_column_letter(jj + 1)].width = 5
                        ws.row_dimensions[ii + 1].height = 30
                        ws.cell(row=ii + 1, column=jj + 1).alignment = Alignment(horizontal='center', vertical='center')
                        if j_ext >= nr_col_ext:
                            ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed', color="FF0000", underline="single")
                        else:
                            ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed', color="0000FF")

    return wb, matriz

# %%
def matrizes_internas(matriz_FR, nr_lin_ext, nr_col_ext, nr_lin_int,  nr_col_int):
    super_matriz_FR = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int))
    eq_paral_internos = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int))
    eq_serie_internos = np.ones((nr_lin_ext, nr_col_ext))
    for i_ext in range(nr_lin_ext):
        for j_ext in range(nr_col_ext):
            rs = i_ext * nr_lin_int
            re = rs + nr_lin_int
            cs = j_ext * nr_col_int
            ce = cs + nr_col_int
            super_matriz_FR[i_ext, j_ext, :, :] = np.array(matriz_FR[rs:re, cs:ce])
            sbmatriz = super_matriz_FR[i_ext, j_ext, :, :]
            eq_paral_internos[i_ext, j_ext, :] = np.sum(sbmatriz, axis=1)
            eq_serie_internos[i_ext, j_ext] = 1 / np.sum(1 / eq_paral_internos[i_ext, j_ext, :])

    return super_matriz_FR, eq_paral_internos, eq_serie_internos

# %%
def calcular_corrente_tensao(V_ao, omega, eq_serie_externos_A1, eq_paral_externos_A1, eq_serie_internos_A1, eq_paral_internos_A1, matriz_A1):
    I_a1 = V_ao * (1j*omega*eq_serie_externos_A1)
    V_a1_ser = I_a1 * 1 / (1j*omega*eq_paral_externos_A1)
    # V_a1_ser = V_a1_ser.reshape(-1, 1) # necessário para transformar em matriz com uma coluna
    I_a1_par = V_a1_ser * (1j*omega*eq_serie_internos_A1)
    V_a1_ser_int = I_a1_par * 1 / (1j*omega*eq_paral_internos_A1)
    I_a1_par_int = V_a1_ser_int * (1j*omega*matriz_A1)
    return I_a1, V_a1_ser, I_a1_par, V_a1_ser_int, I_a1_par_int

# %%
def colorir_planilha_especifica(wb, planilha, nr_lin_int, nr_col_int, nr_lin_ext, nr_col_ext):

    cores = ['FFFFCC', 'CCFFCC', 'CCCCFF', 'FFCCFF']
    ws = wb[planilha]
    for i_ext in range(nr_lin_ext):
        for j_ext in range(2 * nr_col_ext):
            cor = cores[(i_ext + j_ext) % len(cores)]
            fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
            for i_int in range(nr_lin_int):
                for j_int in range(nr_col_int):
                    ii = i_ext * nr_lin_int + i_int
                    jj = j_ext * nr_col_int + j_int
                    ws.row_dimensions[ii + 1].height = 30
                    ws.cell(row=ii + 1, column=jj + 1).alignment = Alignment(horizontal='center', vertical='center')
                    if j_ext >= nr_col_ext:
                        ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed', color="FF0000", underline="single")
                    else:
                        ws.cell(row=ii + 1, column=jj + 1).font = Font(name='Bahnschrift SemiBold Condensed', color="0000FF")

    wb.save('correntes.xlsx')

# %%
def fase(matriz, nr_col_ext, nr_col_int, nr_lin_ext, nr_lin_int, ramo1, ramo2):
    matriz1 = matriz[:, :nr_col_ext * nr_col_int]
    matriz2 = matriz[:, nr_col_ext * nr_col_int:]

    super_matriz1, eq_paral_internos1, eq_serie_internos1 = matrizes_internas(matriz_FR=matriz1, nr_lin_ext=nr_lin_ext, nr_col_ext=nr_col_ext, nr_lin_int=nr_lin_int, nr_col_int=nr_col_int)
    super_matriz2, eq_paral_internos2, eq_serie_internos2 = matrizes_internas(matriz_FR=matriz2, nr_lin_ext=nr_lin_ext, nr_col_ext=nr_col_ext, nr_lin_int=nr_lin_int, nr_col_int=nr_col_int)

    eq_paral_externos1 = (np.sum(eq_serie_internos1, axis=1)).reshape(-1, 1)
    eq_serie_externos1 = 1 / np.sum(1 / eq_paral_externos1)

    eq_paral_externos2 = (np.sum(eq_serie_internos2, axis=1)).reshape(-1, 1)
    eq_serie_externos2 = 1 / np.sum(1 / eq_paral_externos2)

    eq_paral_ramos = eq_serie_externos1 + eq_serie_externos2

    return [eq_paral_ramos, eq_serie_externos1, eq_paral_externos1, eq_serie_internos1, eq_paral_internos1, eq_serie_externos2, eq_paral_externos2, eq_serie_internos2, eq_paral_internos2, matriz1, matriz2, super_matriz1, super_matriz2]
# %%
def calcular_correntes_tensoes(Za, Zb, Zc, Vff, a, matriz_impedancia_sistema):
    matriz_impedancia_malha = np.array([[Za+Zb, -Zb], [-Zb, Zb+Zc]])
    matriz_fontes_malha = np.array([[Vff], [Vff*(a**2)]])
    matriz_correntes_alphabeta = np.linalg.inv(matriz_impedancia_malha) @ matriz_fontes_malha
    I_alpha = matriz_correntes_alphabeta[0, 0]
    I_beta = matriz_correntes_alphabeta[1, 0]
    matriz_correntes_fase = np.array([[I_alpha], [I_beta-I_alpha], [-I_beta]])
    matriz_tensoes_Vabco = matriz_impedancia_sistema @ matriz_correntes_fase
    tensao_deslocamento_netro = np.sum(matriz_tensoes_Vabco) / 3
    return matriz_correntes_fase, matriz_tensoes_Vabco, tensao_deslocamento_netro
