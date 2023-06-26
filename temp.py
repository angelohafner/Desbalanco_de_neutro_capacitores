import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
from engineering_notation import EngNumber
from funcoes_desbalanco_neutro import *



# %%
a = 1*np.exp(1j*2*np.pi/3)
nr_lin_int = 5
nr_col_int = 4
nr_lin_ext = 3
nr_col_ext = 2

potencia_nominal_trifásica = 10e6
omega = 2*np.pi*60
tensao_nominal_fase_fase = 69e3
Vff = tensao_nominal_fase_fase * np.exp(1j*np.pi/6)
tensao_fase_neutro = tensao_nominal_fase_fase / np.sqrt(3)
corrente_fase_neutro = (potencia_nominal_trifásica/3) / tensao_fase_neutro
reatancia = tensao_fase_neutro / corrente_fase_neutro
corrente_nominal_lata = corrente_fase_neutro

cap_total = 1 / (omega*reatancia)
cap_interna = cap_total * (nr_lin_int + nr_lin_ext) / (nr_col_int + nr_col_ext)
des_int = 0.2

print("cap_total = ", cap_total)
print("cap_interna*2 = ", EngNumber(2*cap_interna))

# %% gerar matriz no excel
wb, matriz = matriz_fases_ramos(nr_lin_ext=nr_lin_ext, nr_col_ext=nr_col_ext, nr_lin_int=nr_lin_int, nr_col_int=nr_col_int, cap_interna=cap_interna, des_int=des_int)
wb.save('matriz_total.xlsx')

# %% ler matriz do excel, depois de editada manualmente
data = load_excel_to_numpy('matriz_total.xlsx')
matriz_A = data['A']
matriz_B = data['B']
matriz_C = data['C']


# %%
eq_paral_ramos_A, \
    eq_serie_externos_A1, eq_paral_externos_A1, eq_serie_internos_A1, eq_paral_internos_A1, \
    eq_serie_externos_A2, eq_paral_externos_A2, eq_serie_internos_A2, eq_paral_internos_A2, \
    matriz_A1, matriz_A2, super_matriz_A1, super_matriz_A2 = \
    fase(matriz_A, nr_col_ext, nr_col_int, nr_lin_ext, nr_lin_int, 1, 2)

eq_paral_ramos_B, \
    eq_serie_externos_B1, eq_paral_externos_B1, eq_serie_internos_B1, eq_paral_internos_B1, \
    eq_serie_externos_B2, eq_paral_externos_B2, eq_serie_internos_B2, eq_paral_internos_B2, \
    matriz_B1, matriz_B2, super_matriz_B1, super_matriz_B2 = \
    fase(matriz_B, nr_col_ext, nr_col_int, nr_lin_ext, nr_lin_int, 1, 2)

eq_paral_ramos_C, \
    eq_serie_externos_C1, eq_paral_externos_C1, eq_serie_internos_C1, eq_paral_internos_C1, \
    eq_serie_externos_C2, eq_paral_externos_C2, eq_serie_internos_C2, eq_paral_internos_C2, \
    matriz_C1, matriz_C2, super_matriz_C1, super_matriz_C2 = \
    fase(matriz_C, nr_col_ext, nr_col_int, nr_lin_ext, nr_lin_int, 1, 2)

Za = 1 / (1j*omega*eq_paral_ramos_A)
Zb = 1 / (1j*omega*eq_paral_ramos_B)
Zc = 1 / (1j*omega*eq_paral_ramos_C)

# %%
matriz_impedancia_sistema = np.array([[Za, 0, 0], [0, Zb, 0], [0, 0, Zc]])
matriz_correntes_fase, matriz_tensoes_Vabco, tensao_deslocamento_netro = calcular_correntes_tensoes(Za, Zb, Zc, Vff, a, matriz_impedancia_sistema)

V_ao = matriz_tensoes_Vabco[0]
V_bo = matriz_tensoes_Vabco[1]
V_co = matriz_tensoes_Vabco[2]

I_ao = matriz_correntes_fase[0]
I_bo = matriz_correntes_fase[1]
I_co = matriz_correntes_fase[2]
#
# # %% Começa o caminho inverso
#
# #%% Fase A
# df_Ca_eq_paral_ramos_A = pd.DataFrame([eq_paral_ramos_A])
# df_Ca_eq_ext = pd.DataFrame(np.append(eq_serie_externos_A1, eq_serie_externos_A2))
# df_Ca_pa_ext = pd.DataFrame(np.append(eq_paral_externos_A1, eq_paral_externos_A2, axis=1))
# df_Ca_eq_int = pd.DataFrame(np.append(eq_serie_internos_A1, eq_serie_internos_A2, axis=1))
# df_Ca_pa_int = pd.DataFrame(np.append(eq_paral_internos_A1, eq_paral_internos_A2, axis=1))
# df_Ca__todos = pd.DataFrame(matriz_A)
#
#
# with pd.ExcelWriter('capacitancias_nas_latas.xlsx', engine='openpyxl') as writer:
#     df_Ca_eq_ext.T.to_excel(writer, sheet_name='df_Ca_eq_ext', index=False, header=False)
#     df_Ca_pa_ext.to_excel(writer, sheet_name='df_Ca_pa_ext', index=False, header=False)
#     df_Ca_eq_int.to_excel(writer, sheet_name='df_Ca_eq_int', index=False, header=False)
#     df_Ca_pa_int.to_excel(writer, sheet_name='df_Ca_pa_int', index=False, header=False)
#     df_Ca__todos.to_excel(writer, sheet_name='df_Ca__todos', index=False, header=False)
#

# corrente do capacitor equivalente
I_a1_ext = V_ao * (1j*omega*eq_serie_externos_A1)
# tensoes dos capacitores que estao em serie
V_a1_ser_ext = I_a1_ext / (1j*omega*eq_paral_externos_A1)
# corrente dos capacitores que estao em paralelo
I_a1_par_ext = V_a1_ser_ext * (1j*omega*eq_serie_internos_A1)

# AGORA VAMOS PARA OS INTERNOS
I_a1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
V_a1_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        # capacitores internos em série
        V_a1_ser_int[i_ext, j_ext, :] = I_a1_par_ext[i_ext, j_ext] / (1j * omega * eq_paral_internos_A1[i_ext, j_ext, :])
        I_a1_par_int[i_ext, j_ext, :, :] = V_a1_ser_int[i_ext, j_ext, :].reshape(-1, 1) * (1j*super_matriz_A1[i_ext, j_ext, :, :])

I_a1_par_int_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext*nr_col_int), dtype='complex')
for i_ext in range(nr_lin_ext):
    for j_ext in range(nr_col_ext):
        rs = i_ext * nr_lin_int
        re = rs + nr_lin_int
        cs = j_ext * nr_col_int
        ce = cs + nr_col_int
        I_a1_par_int_excel[rs:re, cs:ce] = I_a1_par_int[i_ext, j_ext, :, :]


df_I_a1_par_int_excel = pd.DataFrame(np.abs(I_a1_par_int_excel))
filename = 'correntes.xlsx'
sheetname = 'internos-a'


with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df_I_a1_par_int_excel.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)

df_I_a1_par_ext = pd.DataFrame(np.abs(I_a1_par_ext))
filename = 'correntes.xlsx'
sheetname = 'externos-a'
with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
    df_I_a1_par_ext.to_excel(writer, sheet_name=sheetname, index=False, header=False)
destaca_maiores_que_nominal(planilha=sheetname, aquivo=filename)









# V_a1_ser_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_ext), dtype='complex')
# I_a1_par_int = np.ones((nr_lin_ext, nr_col_ext, nr_lin_int, nr_col_int), dtype='complex')
# V_a1_ser_int_para_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_ext), dtype='complex')
# I_a1_par_int_para_excel = np.ones((nr_lin_ext*nr_lin_int, nr_col_int), dtype='complex')

# para cada conjunto interno equipavlente paralelo (que está em série)
# for i_ext in range(nr_lin_ext):
#     rs = i_ext * nr_lin_int
#     re = rs + nr_lin_int
#     for j_ext in range(nr_col_ext):
#         V_a1_ser_int[i_ext, j_ext, :, :] = I_a1_par[i_ext, j_ext] * 1 / (1j*omega*eq_paral_internos_A1[rs:re, 0:nr_col_ext])
#         V_a1_ser_int_para_excel[rs:re, 0:nr_col_ext] = I_a1_par[i_ext, j_ext] * 1 / (1j*omega*eq_paral_internos_A1[rs:re, 0:nr_col_ext])
#
# for i_ext in range(nr_lin_ext):
#     rs = i_ext * nr_lin_int
#     re = rs + nr_lin_int
#     I_a1_par_int[i_ext, 0:nr_col_int, :, :] = V_a1_ser_int[i_ext, 0:nr_col_int, :, :] * (1j*omega*super_matriz_A1[rs, re, :, :])
#     # I_a1_par_int_para_excel[rs:re, cs:ce] = V_a1_ser_int[i_ext, j_ext, :, :] * (1j*omega*matriz_A1[rs:re, cs:ce])



# I_a1, V_a1_ser, I_a1_par, V_a1_ser_int, I_a1_par_int = calcular_corrente_tensao(V_ao, omega, eq_serie_externos_A1, eq_paral_externos_A1, eq_serie_internos_A1, eq_paral_internos_A1, matriz_A1)
# I_a2, V_a2_ser, I_a2_par, V_a2_ser_int, I_a2_par_int = calcular_corrente_tensao(V_ao, omega, eq_serie_externos_A2, eq_paral_externos_A2, eq_serie_internos_A2, eq_paral_internos_A2, matriz_A2)
# I_a_par = np.hstack((I_a1_par, I_a2_par))
# df_I_a_par = pd.DataFrame(np.abs(I_a_par))
# I_a_par_int = np.hstack((I_a1_par_int, I_a2_par_int))
# df_I_a_par_int = pd.DataFrame(np.abs(I_a_par_int))
#
# I_b1, V_b1_ser, I_b1_par, V_b1_ser_int, I_b1_par_int = calcular_corrente_tensao(V_ao, omega, eq_serie_externos_B1, eq_paral_externos_B1, eq_serie_internos_B1, eq_paral_internos_B1, matriz_B1)
# I_b2, V_b2_ser, I_b2_par, V_b2_ser_int, I_b2_par_int = calcular_corrente_tensao(V_ao, omega, eq_serie_externos_B2, eq_paral_externos_B2, eq_serie_internos_B2, eq_paral_internos_B2, matriz_B2)
# I_b_par = np.hstack((I_b1_par, I_b2_par))
# df_I_b_par = pd.DataFrame(np.abs(I_b_par))
# I_b_par_int = np.hstack((I_b1_par_int, I_b2_par_int))
# df_I_b_par_int = pd.DataFrame(np.abs(I_b_par_int))
#
# I_c1, V_c1_ser, I_c1_par, V_c1_ser_int, I_c1_par_int = calcular_corrente_tensao(V_ao, omega, eq_serie_externos_C1, eq_paral_externos_C1, eq_serie_internos_C1, eq_paral_internos_C1, matriz_C1)
# I_c2, V_c2_ser, I_c2_par, V_c2_ser_int, I_c2_par_int = calcular_corrente_tensao(V_ao, omega, eq_serie_externos_C2, eq_paral_externos_C2, eq_serie_internos_C2, eq_paral_internos_C2, matriz_C2)
# I_c_par = np.hstack((I_c1_par, I_c2_par))
# df_I_c_par = pd.DataFrame(np.abs(I_c_par))
# I_c_par_int = np.hstack((I_c1_par_int, I_c2_par_int))
# df_I_c_par_int = pd.DataFrame(np.abs(I_c_par_int))
#
# with pd.ExcelWriter('correntes.xlsx', engine='openpyxl') as writer:
#     df_I_a_par.to_excel(writer, sheet_name='df_I_a_par', index=False, header=False)
#     df_I_b_par.to_excel(writer, sheet_name='df_I_b_par', index=False, header=False)
#     df_I_c_par.to_excel(writer, sheet_name='df_I_c_par', index=False, header=False)
#     df_I_a_par_int.to_excel(writer, sheet_name='df_I_a_par_int', index=False, header=False)
#     df_I_b_par_int.to_excel(writer, sheet_name='df_I_b_par_int', index=False, header=False)
#     df_I_c_par_int.to_excel(writer, sheet_name='df_I_c_par_int', index=False, header=False)
#
# destaca_maiores_que_nominal(planilha='df_I_a_par', aquivo='correntes.xlsx', valor_nominal=corrente_nominal_lata)
# destaca_maiores_que_nominal(planilha='df_I_b_par', aquivo='correntes.xlsx', valor_nominal=corrente_nominal_lata)
# destaca_maiores_que_nominal(planilha='df_I_c_par', aquivo='correntes.xlsx', valor_nominal=corrente_nominal_lata)
#
#
# wb = load_workbook('correntes.xlsx')
# colorir_planilha_especifica(wb, 'df_I_a_par_int', nr_lin_int, nr_col_int, nr_lin_ext, nr_col_ext)
# colorir_planilha_especifica(wb, 'df_I_b_par_int', nr_lin_int, nr_col_int, nr_lin_ext, nr_col_ext)
# colorir_planilha_especifica(wb, 'df_I_c_par_int', nr_lin_int, nr_col_int, nr_lin_ext, nr_col_ext)
#
#
#
# # %%
#
